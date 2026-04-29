"""
Genera tests/fixtures/emis_sample_50.xlsx con las primeras 50 empresas del
export real de EMIS — preservando headers en fila 8 y notas en filas 5-6
para que el integration test refleje la estructura real del archivo.

Idempotente: re-ejecutar produce un archivo byte-aproximado idéntico.

Uso:
    python scripts/build_emis_fixture.py
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCE = REPO_ROOT / "data" / "raw" / "emis" / "EMIS_export_2026-04-29.xlsx"
TARGET = REPO_ROOT / "tests" / "fixtures" / "emis_sample_50.xlsx"
N_COMPANIES = 50


def build_fixture() -> None:
    if not SOURCE.exists():
        print(f"ERROR: no existe {SOURCE}", file=sys.stderr)
        sys.exit(1)

    src = load_workbook(SOURCE, data_only=True)
    src_ws = src["Worksheet"]

    out = Workbook()
    assert out.active is not None
    out_ws = out.active
    out_ws.title = "Worksheet"

    # Copiar filas 1-8 (preámbulo + headers) tal cual
    for row_idx in range(1, 9):
        for col_idx, cell in enumerate(src_ws[row_idx], start=1):
            out_ws.cell(row=row_idx, column=col_idx, value=cell.value)

    # Copiar filas 9 .. 9 + N_COMPANIES - 1 (datos)
    for row_idx in range(9, 9 + N_COMPANIES):
        for col_idx, cell in enumerate(src_ws[row_idx], start=1):
            out_ws.cell(row=row_idx, column=col_idx, value=cell.value)

    TARGET.parent.mkdir(parents=True, exist_ok=True)
    out.save(TARGET)
    print(f"✓ {TARGET} ({N_COMPANIES} empresas)")


if __name__ == "__main__":
    build_fixture()
