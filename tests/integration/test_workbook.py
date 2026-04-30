"""
Integration: pipeline completo genera un workbook abrible y bien estructurado.

Cubre:
- Las 11 hojas existen.
- Cada hoja tiene contenido (max_row > 0, max_column > 0).
- No hay celdas con #REF! o #VALUE! evidentes.
- Idempotencia: dos generaciones consecutivas producen valores idénticos
  (se compara cell-by-cell de Pipeline PE como muestra).
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from quimibond.config_loader import load_config
from quimibond.ingestion import EmisLoader
from quimibond.output import generate_workbook
from quimibond.pe_classification import classify_universe

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
SAMPLE = REPO_ROOT / "tests" / "fixtures" / "emis_sample_50.xlsx"
CONFIG_DIR = REPO_ROOT / "config"

EXPECTED_SHEETS = (
    "1. Inicio",
    "2. Universo Raw",
    "3. Pipeline PE",
    "4. Tres Palancas",
    "5. Multiple Arbitrage",
    "6. Owner Fatigue",
    "7. Saturation Check",
    "8. Vista Familiar",
    "9. Investment Memo",
    "10. Rúbrica PE",
    "11. Fuentes",
)

pytestmark = pytest.mark.integration


@pytest.fixture(scope="module")
def generated_workbook(tmp_path_factory: pytest.TempPathFactory) -> Path:
    config = load_config(CONFIG_DIR)
    raws = EmisLoader().load(SAMPLE, source_as_of=date(2026, 4, 30))
    data = classify_universe(raws, config, today=date(2026, 4, 30))
    out = tmp_path_factory.mktemp("wb") / "Quimibond_test.xlsx"
    generate_workbook(data, config, out)
    return out


def test_workbook_file_created(generated_workbook: Path) -> None:
    assert generated_workbook.exists()
    assert generated_workbook.stat().st_size > 1000  # >1KB


def test_all_sheets_present(generated_workbook: Path) -> None:
    wb = load_workbook(generated_workbook)
    try:
        for sheet in EXPECTED_SHEETS:
            assert sheet in wb.sheetnames, f"falta hoja: {sheet}"
    finally:
        wb.close()


def test_sheets_have_content(generated_workbook: Path) -> None:
    wb = load_workbook(generated_workbook)
    try:
        for sheet in EXPECTED_SHEETS:
            ws = wb[sheet]
            assert ws.max_row > 1, f"{sheet} sin filas"
            assert ws.max_column > 0, f"{sheet} sin columnas"
    finally:
        wb.close()


def test_no_error_strings_in_cells(generated_workbook: Path) -> None:
    """Ninguna celda debe contener literales de error tipo #REF! / #VALUE!."""
    wb = load_workbook(generated_workbook)
    error_markers = ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#NULL!")
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if isinstance(v, str):
                        for marker in error_markers:
                            assert marker not in v, f"{sheet}: error literal '{marker}' en celda"
    finally:
        wb.close()


def test_inicio_has_total_companies(generated_workbook: Path) -> None:
    wb = load_workbook(generated_workbook)
    try:
        ws = wb["1. Inicio"]
        # Algún valor de la hoja debe ser 50 (n_companies del sample).
        found = False
        for row in ws.iter_rows(values_only=True):
            if 50 in (v for v in row if isinstance(v, int)):
                found = True
                break
        assert found, "esperaba ver 50 en hoja Inicio"
    finally:
        wb.close()


def test_pipeline_pe_has_role_sections(generated_workbook: Path) -> None:
    """La hoja debe contener al menos un rol operacional (no STRATEGIC)."""
    wb = load_workbook(generated_workbook)
    try:
        ws = wb["3. Pipeline PE"]
        all_text = " ".join(
            str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None
        )
        operational_roles = ("PLATFORM_CANDIDATE", "PRIMARY_BOLT_ON", "TUCK_IN", "UNKNOWN_FIT")
        assert any(token in all_text for token in operational_roles), (
            f"Pipeline PE debe contener al menos un rol operacional. "
            f"Vistos en hoja: {[t for t in operational_roles if t in all_text]}"
        )
    finally:
        wb.close()


def test_multiple_arbitrage_has_formulas(generated_workbook: Path) -> None:
    """Las celdas de MOIC/IRR deben ser fórmulas Excel, no valores estáticos."""
    wb = load_workbook(generated_workbook)
    try:
        ws = wb["5. Multiple Arbitrage"]
        # Buscar al menos una fórmula `=` en la hoja
        n_formulas = 0
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and v.startswith("="):
                    n_formulas += 1
        assert n_formulas > 0, "Multiple Arbitrage debe tener fórmulas Excel"
    finally:
        wb.close()


def test_idempotent_generation(tmp_path: Path) -> None:
    """Dos generaciones consecutivas producen el mismo set de valores."""
    config = load_config(CONFIG_DIR)
    raws = EmisLoader().load(SAMPLE, source_as_of=date(2026, 4, 30))
    data = classify_universe(raws, config, today=date(2026, 4, 30))

    out1 = tmp_path / "wb1.xlsx"
    out2 = tmp_path / "wb2.xlsx"
    generate_workbook(data, config, out1)
    generate_workbook(data, config, out2)

    wb1 = load_workbook(out1)
    wb2 = load_workbook(out2)
    try:
        # Comparar hoja Pipeline PE celda por celda
        ws1 = wb1["3. Pipeline PE"]
        ws2 = wb2["3. Pipeline PE"]
        assert ws1.max_row == ws2.max_row
        assert ws1.max_column == ws2.max_column
        for r in range(1, ws1.max_row + 1):
            for c in range(1, ws1.max_column + 1):
                v1 = ws1.cell(row=r, column=c).value
                v2 = ws2.cell(row=r, column=c).value
                assert v1 == v2, f"diferencia en celda ({r},{c}): {v1!r} vs {v2!r}"
    finally:
        wb1.close()
        wb2.close()
