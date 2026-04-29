"""
Integration tests del EmisLoader contra el fixture sample (50 empresas reales).

Si EMIS cambia formato, estos tests son los primeros en romperse.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest

from quimibond.ingestion import EmisLoader, SourceLoader

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
SAMPLE = REPO_ROOT / "tests" / "fixtures" / "emis_sample_50.xlsx"


pytestmark = pytest.mark.integration


@pytest.fixture(scope="module")
def loader() -> EmisLoader:
    return EmisLoader()


@pytest.fixture(scope="module")
def companies(loader: EmisLoader) -> tuple:
    return loader.load(SAMPLE, source_as_of=date(2026, 4, 29))


def test_loader_implements_protocol() -> None:
    assert isinstance(EmisLoader(), SourceLoader)


def test_loads_50_companies(companies: tuple) -> None:
    assert len(companies) == 50, f"esperaba 50 empresas, obtuve {len(companies)}"


def test_first_company_is_kaltex(companies: tuple) -> None:
    """Sanity: la fila 9 del export real es Kaltex."""
    first = companies[0]
    assert first.source == "EMIS"
    assert first.source_as_of == date(2026, 4, 29)
    assert "Kaltex Textiles" in first.company_name
    assert first.country == "Mexico"
    assert first.source_id == "8340045"
    assert first.rfc == "KTE8207087I9"
    assert first.revenue_usd_mm == 818.09
    assert first.employees == 3500
    assert first.incorporation_year == 1982


def test_naics_extracted_from_main_activities(companies: tuple) -> None:
    """Kaltex tiene Main Activities (NAICS) = 'Broadwoven Fabric Mills(31321)'."""
    kaltex = companies[0]
    assert kaltex.naics == "31321"


def test_all_have_emis_id_as_source_id(companies: tuple) -> None:
    for c in companies:
        assert c.source_id, f"empresa sin source_id: {c.company_name}"
        assert c.source_id.isdigit() or c.source_id.replace("-", "").isdigit()


def test_all_have_company_name(companies: tuple) -> None:
    for c in companies:
        assert c.company_name.strip(), f"empresa con nombre vacío: {c.source_id}"


def test_all_companies_in_mexico(companies: tuple) -> None:
    for c in companies:
        assert c.country == "Mexico", (c.source_id, c.country)


def test_revenue_in_reasonable_range(companies: tuple) -> None:
    """Revenue en USD millones — todos < 5000 (no hay textileras > $5B)."""
    for c in companies:
        if c.revenue_usd_mm is None:
            continue
        assert 0 <= c.revenue_usd_mm < 5000, (c.source_id, c.revenue_usd_mm)


def test_employees_non_negative(companies: tuple) -> None:
    for c in companies:
        if c.employees is not None:
            assert c.employees >= 0


def test_extras_carries_emis_specific_fields(companies: tuple) -> None:
    kaltex = companies[0]
    assert "industry_emis" in kaltex.extra
    assert "all_naics" in kaltex.extra
    # Kaltex debe tener al menos un NAICS de holding y uno textil
    naics_codes = kaltex.extra["all_naics"]
    assert "31321" in naics_codes


def test_rejects_missing_file(tmp_path: Path, loader: EmisLoader) -> None:
    with pytest.raises(FileNotFoundError):
        loader.load(tmp_path / "no_existe.xlsx", source_as_of=date(2026, 4, 29))


def test_rejects_bad_headers(tmp_path: Path, loader: EmisLoader) -> None:
    """Si los headers cambian de fila u orden, debe fallar limpio."""
    from openpyxl import Workbook

    bad = tmp_path / "bad.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Worksheet"
    ws["A8"] = "Wrong"
    ws["B8"] = "Headers"
    wb.save(bad)

    with pytest.raises(ValueError, match="faltan headers requeridos"):
        loader.load(bad, source_as_of=date(2026, 4, 29))
