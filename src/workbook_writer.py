"""
Escribe las empresas nuevas en una copia del workbook,
preservando la estructura, formato y fórmulas existentes.
"""

from __future__ import annotations

import logging
import os
import shutil
from pathlib import Path

import click
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src import config

logger = logging.getLogger(__name__)

THIN = Side(border_style="thin", color="BFBFBF")
LIGHT_GREY = "F2F2F2"
GREEN_GOOD = "C6EFCE"
ORANGE_WARN = "FFD966"


def aplicar_estilo_celda(cell, fila_par: bool, columna: int, status: str | None) -> None:
    """Aplica el mismo estilo que las filas existentes."""
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(
        horizontal="left" if columna in [2, 3, 4, 5, 12, 13] else "center",
        vertical="center",
        wrap_text=True,
        indent=1 if columna in [2, 3, 4, 5, 12, 13] else 0,
    )
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    if fila_par:
        cell.fill = PatternFill("solid", start_color=LIGHT_GREY)
    # Color especial en columna Status (col 14)
    if columna == 14 and status:
        if status == "Prioridad Alta":
            cell.fill = PatternFill("solid", start_color=GREEN_GOOD)
            cell.font = Font(name="Arial", size=10, bold=True)
        elif status == "Investigar":
            cell.fill = PatternFill("solid", start_color=ORANGE_WARN)
        elif status == "Baja prioridad":
            cell.fill = PatternFill("solid", start_color=LIGHT_GREY)


def escribir_filas(workbook_path: Path, nuevas: pd.DataFrame, output_path: Path) -> None:
    """Crea copia del workbook y agrega las filas nuevas en hoja '2. Universo'."""
    if nuevas.empty:
        logger.warning("No hay filas nuevas para escribir.")
        return

    # Copiar primero, luego abrir la copia
    shutil.copy(workbook_path, output_path)
    wb = load_workbook(output_path)
    ws = wb["2. Universo"]

    # Encontrar la siguiente fila libre
    fila_inicio = 5
    while ws.cell(row=fila_inicio, column=2).value:
        fila_inicio += 1
    logger.info("Insertando desde fila %d", fila_inicio)

    # Las primeras 14 columnas son las del esquema original;
    # el resto son los datos adicionales DENUE/Veritrade que el workbook
    # original no tenía. Los escribimos también, ampliando la tabla.
    for offset, (_, row) in enumerate(nuevas.iterrows()):
        fila = fila_inicio + offset
        fila_par = fila % 2 == 0
        status = row.get("Status")
        for col_idx, col_name in enumerate(config.COLUMNAS_FINAL, start=1):
            valor = row.get(col_name)
            if pd.isna(valor):
                valor = None
            cell = ws.cell(row=fila, column=col_idx, value=valor)
            aplicar_estilo_celda(cell, fila_par, col_idx, status if isinstance(status, str) else None)
        ws.row_dimensions[fila].height = 28

    # Si las columnas adicionales (15+) no estaban en el header original, las agregamos
    headers_extra = config.COLUMNAS_FINAL[14:]
    for i, h in enumerate(headers_extra, start=15):
        existing = ws.cell(row=4, column=i).value
        if existing != h:
            cell = ws.cell(row=4, column=i, value=h)
            cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="1F3864")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            # Ancho razonable
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(i)].width = 16

    wb.save(output_path)
    logger.info("✓ Workbook guardado en %s", output_path)
    logger.info("  Filas nuevas agregadas: %d", len(nuevas))


@click.command()
@click.option(
    "--input",
    "input_path",
    type=click.Path(exists=True, path_type=Path),
    default=None,
    help="CSV de empresas nuevas (default: output/targets_consolidado.csv).",
)
@click.option(
    "--workbook",
    "workbook_path",
    type=click.Path(exists=True, path_type=Path),
    default=None,
    help="Workbook origen (default: WORKBOOK_PATH del .env).",
)
@click.option(
    "--output",
    "output_path",
    type=click.Path(path_type=Path),
    default=None,
    help="Workbook destino (default: output/Pipeline_MA_Textil_Mexico_Quimibond_v2.xlsx).",
)
@click.option("-v", "--verbose", is_flag=True)
def main(
    input_path: Path | None,
    workbook_path: Path | None,
    output_path: Path | None,
    verbose: bool,
) -> None:
    """Escribe las nuevas empresas en una copia del workbook."""
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(asctime)s %(levelname)s %(message)s")

    if input_path is None:
        input_path = config.OUTPUT_DIR / "targets_consolidado.csv"
        if not input_path.exists():
            raise click.ClickException(
                "No existe output/targets_consolidado.csv. "
                "Corre 'python -m src.enrichment' primero."
            )

    if workbook_path is None:
        env_path = os.getenv("WORKBOOK_PATH", "data/Pipeline_MA_Textil_Mexico_Quimibond.xlsx")
        workbook_path = Path(env_path)
        if not workbook_path.is_absolute():
            workbook_path = config.ROOT / workbook_path

    if output_path is None:
        output_path = config.OUTPUT_DIR / "Pipeline_MA_Textil_Mexico_Quimibond_v2.xlsx"

    nuevas = pd.read_csv(input_path, encoding=config.ENCODING)
    logger.info("Leídas %d empresas nuevas de %s", len(nuevas), input_path)

    escribir_filas(workbook_path, nuevas, output_path)


if __name__ == "__main__":
    main()
