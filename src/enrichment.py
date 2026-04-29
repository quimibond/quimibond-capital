"""
Enriquecimiento: cruza el output limpio de DENUE con el universo
existente del workbook para identificar empresas nuevas vs. ya conocidas.
"""

from __future__ import annotations

import logging
import os
from pathlib import Path

import click
import pandas as pd
from openpyxl import load_workbook
from rapidfuzz import fuzz

from src import config
from src.cleaning import clave_dedup

logger = logging.getLogger(__name__)


def cargar_universo_actual(workbook_path: Path) -> pd.DataFrame:
    """Lee la hoja '2. Universo' del workbook actual."""
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    if "2. Universo" not in wb.sheetnames:
        raise ValueError(f"No existe hoja '2. Universo' en {workbook_path}")

    ws = wb["2. Universo"]
    # Header está en fila 4 (1-indexed)
    headers = [c.value for c in ws[4]]
    rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[1] is None:  # columna B = Empresa, si está vacía hemos terminado
            break
        rows.append(row)
    df = pd.DataFrame(rows, columns=headers)
    wb.close()
    logger.info("Universo actual: %d empresas", len(df))
    return df


def identificar_nuevas(df_denue: pd.DataFrame, df_actual: pd.DataFrame) -> pd.DataFrame:
    """
    Devuelve el subset de df_denue que NO está en df_actual.
    Match por clave_dedup (fuzzy >= threshold).
    """
    if df_actual.empty:
        return df_denue

    claves_actuales = set(df_actual["Empresa"].dropna().apply(clave_dedup))

    def es_nueva(empresa: str) -> bool:
        clave = clave_dedup(empresa)
        if clave in claves_actuales:
            return False
        # Fuzzy contra todas
        for c in claves_actuales:
            if fuzz.ratio(clave, c) >= config.FUZZY_MATCH_THRESHOLD:
                return False
        return True

    mask = df_denue["Empresa"].apply(es_nueva)
    nuevas = df_denue[mask].copy()
    logger.info(
        "De %d empresas DENUE, %d son nuevas (no estaban en el universo).",
        len(df_denue), len(nuevas),
    )
    return nuevas


@click.command()
@click.option(
    "--denue",
    "denue_path",
    type=click.Path(exists=True, path_type=Path),
    default=None,
    help="CSV limpio de DENUE (default: data/clean/denue_clean.csv).",
)
@click.option(
    "--workbook",
    "workbook_path",
    type=click.Path(exists=True, path_type=Path),
    default=None,
    help="Workbook actual (default: WORKBOOK_PATH del .env).",
)
@click.option("-v", "--verbose", is_flag=True)
def main(denue_path: Path | None, workbook_path: Path | None, verbose: bool) -> None:
    """Genera output/targets_consolidado.csv con solo las empresas nuevas."""
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(asctime)s %(levelname)s %(message)s")

    if denue_path is None:
        denue_path = config.CLEAN_DIR / "denue_clean.csv"
        if not denue_path.exists():
            raise click.ClickException(
                "No existe data/clean/denue_clean.csv. Corre 'python -m src.cleaning' primero."
            )

    if workbook_path is None:
        env_path = os.getenv("WORKBOOK_PATH", "data/Pipeline_MA_Textil_Mexico_Quimibond.xlsx")
        workbook_path = Path(env_path)
        if not workbook_path.is_absolute():
            workbook_path = config.ROOT / workbook_path

    if not workbook_path.exists():
        raise click.ClickException(
            f"Workbook no encontrado en {workbook_path}. "
            f"Coloca el archivo ahí o ajusta WORKBOOK_PATH en .env."
        )

    df_denue = pd.read_csv(denue_path, encoding=config.ENCODING)
    df_actual = cargar_universo_actual(workbook_path)
    nuevas = identificar_nuevas(df_denue, df_actual)

    # Renumerar IDs continuando desde el último del universo actual
    if not df_actual.empty and "ID" in df_actual.columns:
        try:
            ultimo_id = int(pd.to_numeric(df_actual["ID"], errors="coerce").max())
        except (ValueError, TypeError):
            ultimo_id = len(df_actual)
    else:
        ultimo_id = 0

    nuevas = nuevas.reset_index(drop=True)
    nuevas["ID"] = nuevas.index + ultimo_id + 1

    output_path = config.OUTPUT_DIR / "targets_consolidado.csv"
    nuevas.to_csv(output_path, index=False, encoding=config.ENCODING)

    logger.info("=" * 60)
    logger.info("Empresas nuevas listas para agregar al workbook: %d", len(nuevas))
    logger.info("Output: %s", output_path)
    logger.info("Siguiente paso: python -m src.workbook_writer")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
