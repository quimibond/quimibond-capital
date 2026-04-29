"""
Loader del export EMIS / ISI Markets.

Estructura del archivo (verificado abril 2026):
- 1 hoja "Worksheet".
- Filas 1-7: preámbulo + notas (revenue/capital en USD millones).
- Fila 8: headers (46 columnas).
- Fila 9 en adelante: datos.

Filtros aplicados:
- Sin Country → fila descartada (ruido del export).
- Operational Status != "Operational" → descartada.
- Sin EMIS ID → descartada (sin clave estable no podemos referenciar la fila).

NO inventamos datos. Cualquier campo no parseable queda como None.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any, Final

import structlog
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from quimibond.ingestion.emis_parsers import (
    parse_all_naics,
    parse_bool_loose,
    parse_csv_list,
    parse_employees,
    parse_naics_first,
    parse_revenue_usd_mm,
    parse_text,
    parse_year,
)
from quimibond.models import RawCompany

log = structlog.get_logger()

SOURCE_NAME: Final[str] = "EMIS"
HEADER_ROW: Final[int] = 8
DATA_START_ROW: Final[int] = 9


# Mapeo header → key interno. Si EMIS cambia headers, basta tocar este dict.
HEADER_MAP: Final[dict[str, str]] = {
    "Num": "num",
    "Country": "country",
    "Company": "company",
    "Industry (EMIS Industries)": "industry_emis",
    "Total Operating Revenue": "revenue",
    "Business Description/Products": "description",
    "Industry (NAICS)": "industry_naics",
    "Main Activities (NAICS)": "main_naics",
    "Secondary Activities (NAICS)": "secondary_naics",
    "Main Activities (EMIS Industries)": "main_emis",
    "Secondary Activities (EMIS Industries)": "secondary_emis",
    "City": "city",
    "State/County": "state",
    "Postal Code": "zip_code",
    "Address": "address",
    "Phone": "phone",
    "Fax": "fax",
    "Email": "email",
    "Website": "website",
    "Social Media": "social_media",
    "Address Type": "address_type",
    "Key Executives": "executives",
    "Import": "import_flag",
    "Export": "export_flag",
    "Incorporation Date": "incorporation",
    "Number of Employees": "employees",
    "Financial Auditors": "auditors",
    "Legal Form": "legal_form",
    "Legal forms, Local": "legal_form_local",
    "Listed/Unlisted": "listed",
    "Operational Status": "operational_status",
    "Shareholders": "shareholders",
    "Subsidiaries": "subsidiaries",
    "Main Products": "main_products",
    "Registered Capital": "registered_capital",
    "Market Share (%)": "market_share",
    "Fiscal Year": "fiscal_year",
    "Audited": "audited",
    "Consolidated": "consolidated",
    "Source": "source_note",
    "EMIS ID": "emis_id",
    "BMV Company Ticker": "bmv_ticker",
    "ISIN": "isin",
    "LEI": "lei",
    "MX-BANCOMEXT": "bancomext",
    "RFC": "rfc",
}


# ---------------------------------------------------------------------------
# Detalles internos
# ---------------------------------------------------------------------------


def _build_index(headers: list[Any]) -> dict[str, int]:
    """Mapea key interno → índice 0-based en la fila."""
    index: dict[str, int] = {}
    for col_idx, raw_header in enumerate(headers):
        if raw_header is None:
            continue
        h = str(raw_header).strip()
        key = HEADER_MAP.get(h)
        if key is not None:
            index[key] = col_idx
    return index


def _validate_required_headers(index: dict[str, int]) -> None:
    required = {"emis_id", "country", "company"}
    missing = required - index.keys()
    if missing:
        raise ValueError(
            f"EMIS export inválido: faltan headers requeridos {sorted(missing)}. "
            f"Verifica que la fila {HEADER_ROW} tenga los nombres de columna esperados."
        )


def _row_is_empty(row: tuple[Any, ...], index: dict[str, int]) -> bool:
    """Una fila es 'vacía' si Country y Company están ausentes."""
    country = row[index["country"]] if index.get("country") is not None else None
    company = row[index["company"]] if index.get("company") is not None else None
    return parse_text(country) is None and parse_text(company) is None


def _build_raw_company(
    row: tuple[Any, ...],
    index: dict[str, int],
    source_as_of: date,
) -> RawCompany | None:
    def get(key: str) -> Any:
        col = index.get(key)
        return row[col] if col is not None else None

    emis_id = parse_text(get("emis_id"))
    if emis_id is None:
        return None

    operational = parse_text(get("operational_status"))
    if operational and operational.lower() != "operational":
        log.debug("emis.skip_non_operational", emis_id=emis_id, status=operational)
        return None

    company_name = parse_text(get("company"))
    if company_name is None:
        log.warning("emis.skip_missing_company_name", emis_id=emis_id)
        return None

    # NAICS: prioridad 'Main Activities' (más específico), fallback al primero
    # de Industry (NAICS) si Main no trae código.
    naics = parse_naics_first(get("main_naics")) or parse_naics_first(get("industry_naics"))

    # Booleanos export/import: presencia de cualquier marca → True.
    is_exporter = parse_bool_loose(get("export_flag"))

    extras: dict[str, Any] = {}
    for k in (
        "industry_emis",
        "industry_naics",
        "main_naics",
        "secondary_naics",
        "main_emis",
        "secondary_emis",
        "social_media",
        "address_type",
        "fax",
        "auditors",
        "legal_form",
        "legal_form_local",
        "listed",
        "operational_status",
        "subsidiaries",
        "main_products",
        "registered_capital",
        "market_share",
        "fiscal_year",
        "audited",
        "consolidated",
        "source_note",
        "bmv_ticker",
        "isin",
        "lei",
        "bancomext",
        "import_flag",
        "export_flag",
    ):
        v = parse_text(get(k))
        if v is not None:
            extras[k] = v

    naics_codes_all = parse_all_naics(get("industry_naics"))
    if naics_codes_all:
        extras["all_naics"] = list(naics_codes_all)

    return RawCompany(
        source_id=emis_id,
        source=SOURCE_NAME,
        source_as_of=source_as_of,
        company_name=company_name,
        legal_name=company_name,  # EMIS no separa nombre comercial vs legal
        rfc=parse_text(get("rfc")),
        naics=naics,
        activity_description=parse_text(get("description")),
        country=parse_text(get("country")),
        state=parse_text(get("state")),
        city=parse_text(get("city")),
        address=parse_text(get("address")),
        zip_code=parse_text(get("zip_code")),
        revenue_usd_mm=parse_revenue_usd_mm(get("revenue")),
        employees=parse_employees(get("employees")),
        incorporation_year=parse_year(get("incorporation")),
        is_exporter=is_exporter,
        export_destinations=parse_csv_list(get("export_flag")),
        shareholders_text=parse_text(get("shareholders")),
        executives_text=parse_text(get("executives")),
        website=parse_text(get("website")),
        email=parse_text(get("email")),
        phone=parse_text(get("phone")),
        extra=extras,
    )


# ---------------------------------------------------------------------------
# Loader público
# ---------------------------------------------------------------------------


class EmisLoader:
    """SourceLoader para exports EMIS / ISI Markets en xlsx."""

    source_name: Final[str] = SOURCE_NAME

    def load(self, path: Path, source_as_of: date) -> tuple[RawCompany, ...]:
        if not path.is_file():
            raise FileNotFoundError(f"EMIS export no encontrado: {path}")

        log.info("emis.load.start", path=str(path), source_as_of=source_as_of.isoformat())
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            ws: Worksheet = wb["Worksheet"]
            rows_iter = ws.iter_rows(values_only=True)

            # Avanzar al header row (fila 8 = índice 7 en 0-based)
            for _ in range(HEADER_ROW - 1):
                next(rows_iter, None)

            headers_row = next(rows_iter, None)
            if headers_row is None:
                raise ValueError(
                    f"EMIS export vacío o sin headers en fila {HEADER_ROW}: {path}"
                )
            index = _build_index(list(headers_row))
            _validate_required_headers(index)

            companies: list[RawCompany] = []
            skipped_empty = 0
            skipped_other = 0
            for row in rows_iter:
                if _row_is_empty(row, index):
                    skipped_empty += 1
                    continue
                rc = _build_raw_company(row, index, source_as_of)
                if rc is None:
                    skipped_other += 1
                    continue
                companies.append(rc)
        finally:
            wb.close()

        log.info(
            "emis.load.done",
            path=str(path),
            total_loaded=len(companies),
            skipped_empty=skipped_empty,
            skipped_filtered=skipped_other,
        )
        return tuple(companies)
