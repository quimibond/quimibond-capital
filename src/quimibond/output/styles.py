"""
Estilos centralizados para el workbook.

Reglas:
- NamedStyles se registran a nivel workbook una sola vez (StyleSet.register).
- Constantes de color se exponen como enums-like sets para usar fuera de
  NamedStyles (background heatmaps, etc.).
- Sin estilos hardcoded en builders de hojas — todo viene de StyleSet.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Final

from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.workbook import Workbook

from quimibond.models import PERoleType, PriorityType

# ---------------------------------------------------------------------------
# Paleta corporativa
# ---------------------------------------------------------------------------

COLOR_NAVY: Final[str] = "1F3864"
COLOR_NAVY_LIGHT: Final[str] = "5B7FAD"
COLOR_GREY_LIGHT: Final[str] = "F2F2F2"
COLOR_GREY_MED: Final[str] = "BFBFBF"
COLOR_WHITE: Final[str] = "FFFFFF"
COLOR_GREEN: Final[str] = "C6EFCE"
COLOR_GREEN_DARK: Final[str] = "63BE7B"
COLOR_AMBER: Final[str] = "FFEB9C"
COLOR_AMBER_DARK: Final[str] = "FFD966"
COLOR_RED: Final[str] = "F8CBAD"
COLOR_RED_DARK: Final[str] = "F4B084"
COLOR_INPUT: Final[str] = "FFF2CC"  # amarillo claro = editable
COLOR_FORMULA: Final[str] = "E7E6E6"  # gris muy claro = calculado

# Color por rol PE (heatmap operativo)
ROLE_COLORS: Final[dict[PERoleType, str]] = {
    "PLATFORM_CANDIDATE": COLOR_GREEN_DARK,
    "PRIMARY_BOLT_ON": COLOR_GREEN,
    "TUCK_IN": COLOR_AMBER,
    "STRATEGIC": COLOR_GREY_MED,
    "UNKNOWN_FIT": COLOR_GREY_LIGHT,
    "OUT_OF_SCOPE": COLOR_RED,
}

# Color por priority (heatmap subsector)
PRIORITY_COLORS: Final[dict[PriorityType, str]] = {
    "Crítica": COLOR_GREEN_DARK,
    "Alta": COLOR_GREEN,
    "Media": COLOR_GREY_LIGHT,
    "Baja": COLOR_AMBER,
    "Excluida": COLOR_RED,
}

# ---------------------------------------------------------------------------
# Bordes reutilizables
# ---------------------------------------------------------------------------

THIN: Final[Side] = Side(border_style="thin", color=COLOR_GREY_MED)
THICK_BOTTOM: Final[Side] = Side(border_style="medium", color=COLOR_NAVY)

ALL_THIN: Final[Border] = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NO_BORDER: Final[Border] = Border()


# ---------------------------------------------------------------------------
# StyleSet
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class StyleSet:
    """
    Conjunto de NamedStyles que se registra al workbook una vez.

    Uso:
        styles = StyleSet.register(wb)
        cell.style = styles.header  # nombre del estilo
    """

    title: str
    subtitle: str
    section_header: str
    table_header: str
    body: str
    body_centered: str
    body_emphasis: str
    body_input: str
    body_formula: str
    currency_usd_mm: str
    integer: str
    percent: str
    score: str

    @classmethod
    def register(cls, wb: Workbook) -> StyleSet:
        """
        Crea (o reusa) las NamedStyles en el workbook. Idempotente.
        """
        styles_to_register = _build_named_styles()
        for ns in styles_to_register:
            if ns.name not in wb.named_styles:
                wb.add_named_style(ns)
        return cls(
            title="qb_title",
            subtitle="qb_subtitle",
            section_header="qb_section_header",
            table_header="qb_table_header",
            body="qb_body",
            body_centered="qb_body_centered",
            body_emphasis="qb_body_emphasis",
            body_input="qb_body_input",
            body_formula="qb_body_formula",
            currency_usd_mm="qb_currency_usd_mm",
            integer="qb_integer",
            percent="qb_percent",
            score="qb_score",
        )


def _build_named_styles() -> list[NamedStyle]:
    base_font_name = "Calibri"
    title = NamedStyle(name="qb_title")
    title.font = Font(name=base_font_name, size=18, bold=True, color=COLOR_NAVY)
    title.alignment = Alignment(horizontal="left", vertical="center")

    subtitle = NamedStyle(name="qb_subtitle")
    subtitle.font = Font(name=base_font_name, size=11, italic=True, color=COLOR_NAVY_LIGHT)
    subtitle.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    section_header = NamedStyle(name="qb_section_header")
    section_header.font = Font(name=base_font_name, size=12, bold=True, color=COLOR_NAVY)
    section_header.alignment = Alignment(horizontal="left", vertical="center")
    section_header.border = Border(bottom=THICK_BOTTOM)

    table_header = NamedStyle(name="qb_table_header")
    table_header.font = Font(name=base_font_name, size=10, bold=True, color=COLOR_WHITE)
    table_header.fill = PatternFill("solid", start_color=COLOR_NAVY)
    table_header.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    table_header.border = ALL_THIN

    body = NamedStyle(name="qb_body")
    body.font = Font(name=base_font_name, size=10)
    body.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    body.border = ALL_THIN

    body_centered = NamedStyle(name="qb_body_centered")
    body_centered.font = Font(name=base_font_name, size=10)
    body_centered.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_centered.border = ALL_THIN

    body_emphasis = NamedStyle(name="qb_body_emphasis")
    body_emphasis.font = Font(name=base_font_name, size=10, bold=True)
    body_emphasis.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    body_emphasis.border = ALL_THIN

    body_input = NamedStyle(name="qb_body_input")
    body_input.font = Font(name=base_font_name, size=10, bold=True)
    body_input.fill = PatternFill("solid", start_color=COLOR_INPUT)
    body_input.alignment = Alignment(horizontal="center", vertical="center")
    body_input.border = ALL_THIN

    body_formula = NamedStyle(name="qb_body_formula")
    body_formula.font = Font(name=base_font_name, size=10, italic=True)
    body_formula.fill = PatternFill("solid", start_color=COLOR_FORMULA)
    body_formula.alignment = Alignment(horizontal="center", vertical="center")
    body_formula.border = ALL_THIN

    currency = NamedStyle(name="qb_currency_usd_mm")
    currency.font = Font(name=base_font_name, size=10)
    currency.alignment = Alignment(horizontal="right", vertical="center")
    currency.number_format = '_-$#,##0.0,,"M"_-;[Red]_-$#,##0.0,,"M"_-;_-$"-"_-'
    currency.border = ALL_THIN

    integer = NamedStyle(name="qb_integer")
    integer.font = Font(name=base_font_name, size=10)
    integer.alignment = Alignment(horizontal="right", vertical="center")
    integer.number_format = "#,##0"
    integer.border = ALL_THIN

    percent = NamedStyle(name="qb_percent")
    percent.font = Font(name=base_font_name, size=10)
    percent.alignment = Alignment(horizontal="right", vertical="center")
    percent.number_format = "0.0%"
    percent.border = ALL_THIN

    score = NamedStyle(name="qb_score")
    score.font = Font(name=base_font_name, size=10, bold=True)
    score.alignment = Alignment(horizontal="center", vertical="center")
    score.number_format = "0.00"
    score.border = ALL_THIN

    return [
        title,
        subtitle,
        section_header,
        table_header,
        body,
        body_centered,
        body_emphasis,
        body_input,
        body_formula,
        currency,
        integer,
        percent,
        score,
    ]


# ---------------------------------------------------------------------------
# Helpers de coloreado por rol / priority (no son NamedStyles porque varían)
# ---------------------------------------------------------------------------


def role_fill(role: PERoleType) -> PatternFill:
    return PatternFill("solid", start_color=ROLE_COLORS[role])


def priority_fill(priority: PriorityType) -> PatternFill:
    return PatternFill("solid", start_color=PRIORITY_COLORS[priority])


def score_heatmap_fill(score: float) -> PatternFill:
    """Verde si >= 0.70, ámbar si >= 0.50, rojo si < 0.50."""
    if score >= 0.70:
        return PatternFill("solid", start_color=COLOR_GREEN)
    if score >= 0.50:
        return PatternFill("solid", start_color=COLOR_AMBER)
    return PatternFill("solid", start_color=COLOR_RED)
