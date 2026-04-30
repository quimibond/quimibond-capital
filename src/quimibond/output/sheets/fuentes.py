"""Hoja 11 — Fuentes: roadmap multi-fuente del pipeline."""

from __future__ import annotations

from collections import Counter

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from quimibond.config_loader import Config
from quimibond.models import PipelineData
from quimibond.output.helpers import (
    set_column_widths,
    write_section_header,
    write_table_headers,
    write_title,
)
from quimibond.output.styles import StyleSet


SOURCES_ROADMAP: list[tuple[str, str, str, str]] = [
    (
        "EMIS / ISI Markets",
        "Activa",
        "Universo principal: 700 empresas con financieros + accionistas + ejecutivos.",
        "Snapshot xlsx mensual. Cobertura RFC 95%, revenue 30%, empleados 100%.",
    ),
    (
        "DENUE / INEGI",
        "Beta",
        "Censo público de establecimientos. Útil para detectar empresas no en EMIS.",
        "API gratuita. NAICS + estado + estrato. NO devuelve RFC ni financieros.",
    ),
    (
        "Capital IQ vía Tec",
        "Pendiente",
        "Validación de financieros + comparables internacionales.",
        "Requiere acceso institucional Tec. Para diligencia detallada de targets.",
    ),
    (
        "Statista / Euromonitor",
        "Pendiente",
        "Tamaño de mercado por subsegmento + tendencias macro.",
        "Para investment memo: cifras de mercado, crecimiento, share por player.",
    ),
    (
        "Veritrade",
        "Pendiente",
        "Datos de comercio exterior por aduana (HS code, peso, valor, destino).",
        "Cruzar con exporters detectados para validar volúmenes y diversificación.",
    ),
    (
        "CANAINTEX",
        "Manual",
        "Cámara nacional textil. Lista de socios + asambleas para outreach directo.",
        "Datos no estructurados; útil para llamadas en frío y validación de familias.",
    ),
    (
        "SAT / FEA",
        "Pendiente",
        "Cruce de RFCs con avisos de actividad y régimen fiscal.",
        "Para validar que los targets son operativos y al corriente fiscal.",
    ),
]


class FuentesSheet:
    name = "11. Fuentes"

    def build(
        self,
        wb: Workbook,
        ws: Worksheet,
        data: PipelineData,
        config: Config,
        styles: StyleSet,
    ) -> None:
        set_column_widths(ws, [22, 14, 50, 50])

        row = write_title(
            ws, 1,
            "Fuentes — roadmap multi-source",
            "Cómo se construye el universo. Estado actual y siguientes pasos.",
            styles,
        )

        # Distribución actual del universo por source
        source_counts = Counter(c.source for c in data.companies)
        row = write_section_header(ws, row, "Universo actual por fuente", styles)
        row = write_table_headers(ws, row, ["Fuente", "Empresas", "Comentarios", ""], styles)
        for source, count in source_counts.most_common():
            ws.cell(row=row, column=1, value=source).style = styles.body_emphasis
            ws.cell(row=row, column=2, value=count).style = styles.integer
            ws.cell(
                row=row, column=3,
                value="Snapshot único cargado en este run."
            ).style = styles.body
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            row += 1
        row += 1

        # Roadmap
        row = write_section_header(ws, row, "Roadmap de fuentes", styles)
        row = write_table_headers(
            ws, row,
            ["Fuente", "Estado", "Aporte", "Notas técnicas"],
            styles,
        )
        for source, status, aporte, notas in SOURCES_ROADMAP:
            ws.cell(row=row, column=1, value=source).style = styles.body_emphasis
            ws.cell(row=row, column=2, value=status).style = styles.body_centered
            ws.cell(row=row, column=3, value=aporte).style = styles.body
            ws.cell(row=row, column=4, value=notas).style = styles.body
            ws.row_dimensions[row].height = 38
            row += 1
