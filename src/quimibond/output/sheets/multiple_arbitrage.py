"""
Hoja 5 — Multiple Arbitrage.

Modelo LBO simplificado con asunciones EDITABLES en celdas amarillas.
Las columnas calculadas son fórmulas Excel para que cuando Jose cambie
exit_multiple o ebitda_margin se recalcule el output sin re-correr el pipeline.

Layout:
- Filas 1-2: título y descripción.
- Filas 4-7: panel de asunciones editables (ebitda margin, exit multiple,
  hold years, discount rate). Estas son las celdas que el LP/Jose puede mover.
- Fila 9 en adelante: tabla por target con columnas EBITDA, buy multiple,
  EV buy/exit, MOIC, IRR — todas como fórmulas relativas al panel superior.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from quimibond.config_loader import Config
from quimibond.models import PipelineData
from quimibond.output.helpers import (
    freeze_below,
    set_column_widths,
    write_section_header,
    write_table_headers,
    write_title,
)
from quimibond.output.styles import StyleSet

ASSUMPTIONS_ROW_EBITDA = 4
ASSUMPTIONS_ROW_EXIT = 5
ASSUMPTIONS_ROW_HOLD = 6
ASSUMPTIONS_ROW_DISCOUNT = 7
ASSUMPTIONS_VALUE_COL = 2  # B


class MultipleArbitrageSheet:
    name = "5. Multiple Arbitrage"

    def build(
        self,
        wb: Workbook,
        ws: Worksheet,
        data: PipelineData,
        config: Config,
        styles: StyleSet,
    ) -> None:
        set_column_widths(ws, [32, 14, 14, 14, 12, 14, 14, 14, 12, 12])

        write_title(
            ws, 1,
            "Multiple Arbitrage — modelo LBO simplificado",
            "Edita las celdas amarillas para sensibilizar el modelo. Las celdas "
            "grises se recalculan automáticamente al recalcular el workbook.",
            styles,
        )

        # Panel de asunciones (editables)
        playbook = config.pe_playbook
        self._write_assumption(ws, ASSUMPTIONS_ROW_EBITDA,
                               "EBITDA margin asumido",
                               playbook.ebitda_margin_default,
                               styles.percent, styles)
        self._write_assumption(ws, ASSUMPTIONS_ROW_EXIT,
                               "Exit multiple (consolidado)",
                               playbook.exit_multiple_default,
                               styles.body_input, styles)
        self._write_assumption(ws, ASSUMPTIONS_ROW_HOLD,
                               "Hold period (años)",
                               playbook.hold_period_years,
                               styles.body_input, styles)
        self._write_assumption(ws, ASSUMPTIONS_ROW_DISCOUNT,
                               "Tasa de descuento corporativa",
                               playbook.discount_rate,
                               styles.percent, styles)

        # Pintar las celdas de valor de amarillo (input)
        for r in (
            ASSUMPTIONS_ROW_EBITDA,
            ASSUMPTIONS_ROW_EXIT,
            ASSUMPTIONS_ROW_HOLD,
            ASSUMPTIONS_ROW_DISCOUNT,
        ):
            ws.cell(row=r, column=ASSUMPTIONS_VALUE_COL).style = styles.body_input
        # Forzar formatos:
        ws.cell(row=ASSUMPTIONS_ROW_EBITDA, column=ASSUMPTIONS_VALUE_COL).number_format = "0.0%"
        ws.cell(row=ASSUMPTIONS_ROW_DISCOUNT, column=ASSUMPTIONS_VALUE_COL).number_format = "0.0%"

        # Tabla con fórmulas
        targets = [
            c for c in data.companies
            if c.revenue_usd_mm is not None and c.revenue_usd_mm > 0
            and c.pe_role in ("PLATFORM_CANDIDATE", "PRIMARY_BOLT_ON", "TUCK_IN")
        ]
        targets.sort(key=lambda c: -(c.revenue_usd_mm or 0))

        row = 9
        row = write_section_header(ws, row, f"Targets con revenue ({len(targets)})", styles)
        headers = [
            "Empresa",
            "Revenue (USD)",
            "EBITDA est.",
            "Buy multiple",
            "EV buy",
            "EV exit",
            "MOIC",
            "IRR",
            "PE Role",
            "EMIS ID",
        ]
        row = write_table_headers(ws, row, headers, styles)
        freeze_below(ws, row - 1, col=1)

        ebitda_ref = f"$B${ASSUMPTIONS_ROW_EBITDA}"
        exit_ref = f"$B${ASSUMPTIONS_ROW_EXIT}"
        hold_ref = f"$B${ASSUMPTIONS_ROW_HOLD}"

        for c in targets:
            col = 1
            ws.cell(row=row, column=col, value=c.company_name).style = styles.body
            col += 1
            # Revenue en USD (multiplicamos para que el formato custom ',,M' funcione)
            rev_value = (c.revenue_usd_mm or 0.0) * 1_000_000
            rev_letter = get_column_letter(col)
            ws.cell(row=row, column=col, value=rev_value).style = styles.currency_usd_mm
            col += 1

            # EBITDA = Revenue * margin
            ebitda_letter = get_column_letter(col)
            ws.cell(row=row, column=col, value=f"={rev_letter}{row}*{ebitda_ref}").style = (
                styles.currency_usd_mm
            )
            col += 1

            # Buy multiple — pre-calculado del playbook (no fórmula porque
            # depende de bracket lookup; lo dejamos como valor para legibilidad)
            buy_x = c.arbitrage.buy_multiple if c.arbitrage else 0.0
            buy_letter = get_column_letter(col)
            ws.cell(row=row, column=col, value=buy_x).style = styles.body_centered
            ws.cell(row=row, column=col).number_format = "0.0\"x\""
            col += 1

            # EV buy = EBITDA * buy_multiple
            ev_buy_letter = get_column_letter(col)
            ws.cell(
                row=row, column=col,
                value=f"={ebitda_letter}{row}*{buy_letter}{row}",
            ).style = styles.currency_usd_mm
            col += 1

            # EV exit = EBITDA * exit_multiple (panel)
            ev_exit_letter = get_column_letter(col)
            ws.cell(
                row=row, column=col,
                value=f"={ebitda_letter}{row}*{exit_ref}",
            ).style = styles.currency_usd_mm
            col += 1

            # MOIC = EV exit / EV buy
            moic_letter = get_column_letter(col)
            moic_cell = ws.cell(
                row=row, column=col,
                value=f"={ev_exit_letter}{row}/{ev_buy_letter}{row}",
            )
            moic_cell.style = styles.body_centered
            moic_cell.number_format = "0.00\"x\""
            col += 1

            # IRR = MOIC^(1/hold) - 1
            irr_cell = ws.cell(
                row=row, column=col,
                value=f"=({moic_letter}{row})^(1/{hold_ref})-1",
            )
            irr_cell.style = styles.percent
            col += 1

            ws.cell(row=row, column=col, value=c.pe_role).style = styles.body_centered
            col += 1
            ws.cell(row=row, column=col, value=c.emis_id).style = styles.body_centered
            row += 1

    def _write_assumption(
        self,
        ws: Worksheet,
        row: int,
        label: str,
        value: float | int | str,
        value_style: str,  # noqa: ARG002 — applied uniformly below
        styles: StyleSet,
    ) -> None:
        ws.cell(row=row, column=1, value=label).style = styles.body_emphasis
        cell = ws.cell(row=row, column=ASSUMPTIONS_VALUE_COL, value=value)
        cell.style = styles.body_input
