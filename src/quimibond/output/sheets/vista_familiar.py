"""Hoja 8 — Vista Familiar: targets agrupados por familia textil clásica."""

from __future__ import annotations

from collections import defaultdict

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from quimibond.config_loader import Config
from quimibond.models import Company, PipelineData
from quimibond.output.helpers import (
    set_column_widths,
    write_section_header,
    write_table_headers,
    write_title,
)
from quimibond.output.styles import StyleSet


class VistaFamiliarSheet:
    name = "8. Vista Familiar"

    def build(
        self,
        wb: Workbook,
        ws: Worksheet,
        data: PipelineData,
        config: Config,
        styles: StyleSet,
    ) -> None:
        set_column_widths(ws, [22, 11, 32, 16, 22, 14, 12, 11, 50])

        row = write_title(
            ws, 1,
            "Vista Familiar",
            "Targets agrupados por familia textil clásica detectada en accionistas o "
            "ejecutivos. Ideal para conversaciones con papá / abuelo / contactos directos.",
            styles,
        )

        # Bucket por familia
        by_family: dict[str, list[Company]] = defaultdict(list)
        for c in data.companies:
            family = c.classification.detected_family
            if family:
                by_family[family].append(c)

        if not by_family:
            ws.cell(row=row, column=1, value="No se detectaron familias clásicas en este universo.").style = (
                styles.subtitle
            )
            return

        # Ordenar familias por # empresas desc
        ordered_families = sorted(
            by_family.items(),
            key=lambda kv: (-len(kv[1]), kv[0]),
        )

        headers = [
            "Familia",
            "EMIS ID",
            "Empresa",
            "Subsector",
            "PE Role",
            "Revenue",
            "Empleados",
            "Combined",
            "Estado",
        ]
        row = write_section_header(
            ws, row,
            f"{len(ordered_families)} familias · {sum(len(v) for v in by_family.values())} empresas",
            styles,
        )
        row = write_table_headers(ws, row, headers, styles)

        for family, companies in ordered_families:
            companies.sort(key=lambda c: (-c.levers.combined, c.emis_id))
            for i, c in enumerate(companies):
                col = 1
                # Solo escribimos la familia en la primera fila del bloque
                if i == 0:
                    fam_cell = ws.cell(row=row, column=col, value=f"{family} ({len(companies)})")
                    fam_cell.style = styles.body_emphasis
                col += 1
                ws.cell(row=row, column=col, value=c.emis_id).style = styles.body_centered
                col += 1
                ws.cell(row=row, column=col, value=c.company_name).style = styles.body
                col += 1
                ws.cell(row=row, column=col, value=c.classification.subsector).style = styles.body
                col += 1
                ws.cell(row=row, column=col, value=c.pe_role).style = styles.body_centered
                col += 1
                if c.revenue_usd_mm is not None:
                    ws.cell(
                        row=row, column=col,
                        value=c.revenue_usd_mm * 1_000_000,
                    ).style = styles.currency_usd_mm
                else:
                    ws.cell(row=row, column=col, value="—").style = styles.body_centered
                col += 1
                ws.cell(
                    row=row, column=col,
                    value=c.employees if c.employees is not None else "—",
                ).style = styles.body_centered
                col += 1
                ws.cell(row=row, column=col, value=c.levers.combined).style = styles.score
                col += 1
                ws.cell(row=row, column=col, value=c.state or "—").style = styles.body
                row += 1
            row += 1  # espacio entre familias
