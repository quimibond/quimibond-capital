"""Hoja 1 — Inicio: portada + stats agregadas + mapa del workbook."""

from __future__ import annotations

from collections import Counter

from typing import cast

from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from quimibond.config_loader import Config
from quimibond.models import PERoleType, PipelineData, PriorityType
from quimibond.output.helpers import (
    set_column_widths,
    write_kv_pair,
    write_section_header,
    write_title,
)
from quimibond.output.styles import (
    PRIORITY_COLORS,
    StyleSet,
    role_fill,
)


class InicioSheet:
    name = "1. Inicio"

    def build(
        self,
        wb: Workbook,
        ws: Worksheet,
        data: PipelineData,
        config: Config,
        styles: StyleSet,
    ) -> None:
        set_column_widths(ws, [38, 18, 18, 18, 18, 18])

        row = write_title(
            ws, 1,
            "Quimibond Capital — Pipeline PE",
            f"Universo textil México · {data.n_companies} empresas · "
            f"snapshot {data.generated_at.isoformat()}",
            styles,
        )

        # Stats agregadas
        row = write_section_header(ws, row, "Resumen del universo", styles)

        row = write_kv_pair(ws, row, "Total empresas", data.n_companies, styles)
        row = write_kv_pair(
            ws, row,
            "Empresas con revenue declarado",
            sum(1 for c in data.companies if c.revenue_usd_mm is not None),
            styles,
        )
        row = write_kv_pair(
            ws, row,
            "Empresas familiares MX detectadas",
            sum(1 for c in data.companies if c.is_familiar_mx),
            styles,
        )
        row = write_kv_pair(
            ws, row,
            "Subsidiarias extranjeras",
            sum(1 for c in data.companies if c.is_foreign_subsidiary),
            styles,
        )

        # Edad media
        ages = [c.age_years for c in data.companies if c.age_years is not None]
        avg_age = sum(ages) / len(ages) if ages else 0
        row = write_kv_pair(ws, row, "Edad promedio (años)", round(avg_age, 1), styles)

        row += 1

        # Distribución por rol PE
        row = write_section_header(ws, row, "Distribución por rol PE", styles)
        role_counts = Counter(c.pe_role for c in data.companies)
        roles_order: tuple[PERoleType, ...] = (
            "PLATFORM_CANDIDATE",
            "PRIMARY_BOLT_ON",
            "TUCK_IN",
            "STRATEGIC",
            "UNKNOWN_FIT",
            "OUT_OF_SCOPE",
        )
        for role in roles_order:
            cnt = role_counts.get(role, 0)
            label_cell = ws.cell(row=row, column=1, value=role)
            label_cell.style = styles.body_emphasis
            label_cell.fill = role_fill(role)
            ws.cell(row=row, column=2, value=cnt).style = styles.integer
            row += 1

        row += 1

        # Distribución por priority
        row = write_section_header(ws, row, "Distribución por prioridad de subsector", styles)
        prio_counts = Counter(c.classification.subsector_priority for c in data.companies)
        priorities_order: tuple[PriorityType, ...] = ("Crítica", "Alta", "Media", "Baja", "Excluida")
        for prio in priorities_order:
            cnt = prio_counts.get(prio, 0)
            label_cell = ws.cell(row=row, column=1, value=prio)
            label_cell.style = styles.body_emphasis
            label_cell.fill = priority_color_fill(prio)
            ws.cell(row=row, column=2, value=cnt).style = styles.integer
            row += 1

        row += 1

        # Mapa del workbook
        row = write_section_header(ws, row, "Mapa del workbook", styles)
        sheets_map = [
            ("2. Universo Raw", "Las 698 empresas con todos los campos enriquecidos."),
            ("3. Pipeline PE", "Vista operativa: PLATFORM/BOLT_ON/TUCK_IN ordenadas por score."),
            ("4. Tres Palancas", "Detalle por target con justificación de cost/revenue/arbitrage."),
            ("5. Multiple Arbitrage", "Modelo LBO simplificado con asunciones editables."),
            ("6. Owner Fatigue", "Score de fatiga del dueño con razones detectadas."),
            ("7. Saturation Check", "Análisis de saturación por subsegmento."),
            ("8. Vista Familiar", "Targets agrupados por familia textil clásica."),
            ("9. Investment Memo", "Estructura para presentación a LPs externos."),
            ("10. Rúbrica PE", "Playbook documentado: thresholds, múltiplos, pesos."),
            ("11. Fuentes", "Roadmap multi-fuente: EMIS / DENUE / Statista / Capital IQ."),
        ]
        for name, desc in sheets_map:
            ws.cell(row=row, column=1, value=name).style = styles.body_emphasis
            ws.cell(row=row, column=2, value=desc).style = styles.body
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            row += 1


def priority_color_fill(prio: PriorityType) -> PatternFill:
    color = PRIORITY_COLORS.get(prio, "FFFFFF")
    return PatternFill("solid", start_color=color)
