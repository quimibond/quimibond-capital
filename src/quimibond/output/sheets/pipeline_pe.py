"""Hoja 3 — Pipeline PE: la vista operativa principal de Jose."""

from __future__ import annotations

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from quimibond.config_loader import Config
from quimibond.models import Company, PERoleType, PipelineData
from quimibond.output.helpers import (
    freeze_below,
    set_column_widths,
    write_section_header,
    write_table_headers,
    write_title,
)
from quimibond.output.styles import StyleSet, role_fill, score_heatmap_fill

# Solo mostramos los roles operacionales en orden lógico de prioridad
PRIORITY_ROLES: tuple[PERoleType, ...] = (
    "PLATFORM_CANDIDATE",
    "PRIMARY_BOLT_ON",
    "TUCK_IN",
    "UNKNOWN_FIT",
)

HEADERS = [
    "Rank",
    "EMIS ID",
    "Empresa",
    "Subsector",
    "Priority",
    "Estado",
    "Empleados",
    "Revenue (USD MM)",
    "EBITDA est.",
    "Familia",
    "Capital",
    "Cost",
    "Revenue",
    "Arbitrage",
    "Combined",
    "Justificación rol",
]
COL_WIDTHS = [6, 11, 36, 22, 10, 18, 11, 14, 14, 14, 16, 8, 8, 9, 10, 50]


class PipelinePESheet:
    name = "3. Pipeline PE"

    def build(
        self,
        wb: Workbook,
        ws: Worksheet,
        data: PipelineData,
        config: Config,
        styles: StyleSet,
    ) -> None:
        set_column_widths(ws, COL_WIDTHS)

        row = write_title(
            ws, 1,
            "Pipeline PE — vista operativa",
            "Targets accionables ordenados por combined score dentro de cada rol. "
            "STRATEGIC y OUT_OF_SCOPE excluidos (ver hoja 2 si necesitas verlos).",
            styles,
        )

        for role in PRIORITY_ROLES:
            companies_in_role = [c for c in data.companies if c.pe_role == role]
            if not companies_in_role:
                continue
            companies_in_role.sort(key=lambda c: -c.levers.combined)

            row = write_section_header(
                ws, row, f"{role}  ·  {len(companies_in_role)} empresas", styles,
            )
            row = write_table_headers(ws, row, HEADERS, styles)
            if role == PRIORITY_ROLES[0]:
                freeze_below(ws, row - 1, col=3)

            for rank, c in enumerate(companies_in_role, start=1):
                self._write_row(ws, row, rank, c, styles)
                row += 1

            row += 1  # espacio entre secciones

    def _write_row(
        self,
        ws: Worksheet,
        row: int,
        rank: int,
        c: Company,
        styles: StyleSet,
    ) -> None:
        clf = c.classification
        col = 1
        ws.cell(row=row, column=col, value=rank).style = styles.body_centered
        col += 1
        ws.cell(row=row, column=col, value=c.emis_id).style = styles.body_centered
        col += 1
        ws.cell(row=row, column=col, value=c.company_name).style = styles.body
        col += 1
        ws.cell(row=row, column=col, value=clf.subsector).style = styles.body
        col += 1
        ws.cell(row=row, column=col, value=clf.subsector_priority).style = styles.body_centered
        col += 1
        ws.cell(row=row, column=col, value=c.state or "—").style = styles.body
        col += 1

        if c.employees is not None:
            ws.cell(row=row, column=col, value=c.employees).style = styles.integer
        else:
            ws.cell(row=row, column=col, value="—").style = styles.body_centered
        col += 1

        if c.revenue_usd_mm is not None:
            ws.cell(row=row, column=col, value=c.revenue_usd_mm * 1_000_000).style = (
                styles.currency_usd_mm
            )
        else:
            ws.cell(row=row, column=col, value="—").style = styles.body_centered
        col += 1

        if c.ebitda_estimate_usd_mm is not None:
            ws.cell(row=row, column=col, value=c.ebitda_estimate_usd_mm * 1_000_000).style = (
                styles.currency_usd_mm
            )
        else:
            ws.cell(row=row, column=col, value="—").style = styles.body_centered
        col += 1

        ws.cell(row=row, column=col, value=clf.detected_family or "—").style = styles.body_centered
        col += 1
        ws.cell(row=row, column=col, value=clf.capital_origin).style = styles.body_centered
        col += 1

        # Levers
        ws.cell(row=row, column=col, value=c.levers.cost).style = styles.score
        col += 1
        ws.cell(row=row, column=col, value=c.levers.revenue).style = styles.score
        col += 1
        ws.cell(row=row, column=col, value=c.levers.arbitrage).style = styles.score
        col += 1

        combined_cell = ws.cell(row=row, column=col, value=c.levers.combined)
        combined_cell.style = styles.score
        combined_cell.fill = score_heatmap_fill(c.levers.combined)
        col += 1

        ws.cell(row=row, column=col, value=c.pe_role_justification).style = styles.body
        # Recoloreo del rol via celda 5 (Priority) -> en su lugar pintamos rank cell
        ws.cell(row=row, column=1).fill = role_fill(c.pe_role)
