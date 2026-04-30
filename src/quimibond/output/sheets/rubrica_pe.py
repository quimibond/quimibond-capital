"""Hoja 10 — Rúbrica PE: playbook documentado (thresholds, múltiplos, pesos)."""

from __future__ import annotations

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from quimibond.config_loader import Config
from quimibond.models import PipelineData
from quimibond.output.helpers import (
    set_column_widths,
    write_section_header,
    write_table_headers,
    write_title,
)
from quimibond.output.styles import StyleSet


class RubricaPESheet:
    name = "10. Rúbrica PE"

    def build(
        self,
        wb: Workbook,
        ws: Worksheet,
        data: PipelineData,
        config: Config,
        styles: StyleSet,
    ) -> None:
        set_column_widths(ws, [38, 22, 60])

        row = write_title(
            ws, 1,
            "Rúbrica PE — playbook documentado",
            "Thresholds, múltiplos y pesos editables vía config/*.yaml. "
            "Esta hoja refleja los valores activos en este snapshot.",
            styles,
        )

        thresholds = config.thresholds
        playbook = config.pe_playbook

        # Sección 1: Brackets de revenue
        row = write_section_header(ws, row, "Brackets de revenue (USD MM)", styles)
        row = write_table_headers(ws, row, ["Bracket", "Valor", "Significado"], styles)
        rb = thresholds.revenue_brackets_usd_mm
        rb_rows: list[tuple[str, str, str]] = [
            ("platform_min", f"${rb.platform_min:.1f}M", "Mínimo para PLATFORM_CANDIDATE"),
            ("bolt_on_max", f"${rb.bolt_on_max:.1f}M", "Tope superior bolt-on"),
            ("bolt_on_min", f"${rb.bolt_on_min:.1f}M", "Mínimo bolt-on"),
            ("tuck_in_max", f"${rb.tuck_in_max:.1f}M", "Tope tuck-in"),
            ("tuck_in_min", f"${rb.tuck_in_min:.1f}M", "Mínimo tuck-in"),
            ("out_of_scope_max", f"${rb.out_of_scope_max:.1f}M", "< este = OUT_OF_SCOPE"),
        ]
        for label, val_str, sig in rb_rows:
            row = self._row(ws, row, label, val_str, sig, styles)
        row += 1

        # Sección 2: Múltiplos de compra
        row = write_section_header(ws, row, "Múltiplos de compra por tamaño", styles)
        row = write_table_headers(ws, row, ["Bracket", "Buy multiple", "Aplicable a"], styles)
        prev = 0.0
        for b in playbook.buy_multiples_by_size:
            if b.max_revenue is None:
                applicable = f"Revenue > ${prev:.0f}M"
                bracket = "Catch-all"
            else:
                applicable = f"Revenue ${prev:.0f}M – ${b.max_revenue:.0f}M"
                bracket = applicable
                prev = b.max_revenue
            row = self._row(ws, row, bracket, f"{b.multiple:.1f}×", applicable, styles)
        row += 1

        # Sección 3: Asunciones financieras
        row = write_section_header(ws, row, "Asunciones financieras", styles)
        row = write_table_headers(ws, row, ["Variable", "Valor", "Notas"], styles)
        assump_rows: list[tuple[str, str, str]] = [
            ("Exit multiple objetivo", f"{playbook.exit_multiple_default:.1f}×",
             "Múltiplo del consolidado al salir"),
            ("EBITDA margin default", f"{playbook.ebitda_margin_default*100:.0f}%",
             "Margen EBITDA promedio textil mid-market MX"),
            ("Hold period", f"{playbook.hold_period_years} años", "Plazo de tenencia"),
            ("Discount rate", f"{playbook.discount_rate*100:.0f}%",
             "Tasa corporativa Quimibond"),
            ("Tipo de cambio", f"{playbook.exchange_rate_mxn_usd:.0f} MXN/USD",
             "Para conversión MXN→USD"),
        ]
        for label, val_str, sig in assump_rows:
            row = self._row(ws, row, label, val_str, sig, styles)
        row += 1

        # Sección 4: Pesos del score combinado
        row = write_section_header(ws, row, "Pesos del score combinado", styles)
        row = write_table_headers(ws, row, ["Palanca", "Peso", "Tesis"], styles)
        weights = playbook.scoring_weights
        for label, val, sig in [
            ("Cost", f"{weights.lever_cost*100:.0f}%", "Captura de sinergias operativas"),
            ("Revenue", f"{weights.lever_revenue*100:.0f}%", "Cross-sell vía clientes Tier-1"),
            ("Arbitrage", f"{weights.lever_arbitrage*100:.0f}%", "Re-rating múltiplo al consolidar"),
        ]:
            row = self._row(ws, row, label, val, sig, styles)
        row += 1

        # Sección 5: Caps por rol PE
        row = write_section_header(ws, row, "Caps de combined score por rol", styles)
        row = write_table_headers(ws, row, ["Rol", "Cap", "Justificación"], styles)
        for role, cap in playbook.role_combined_cap.items():
            sig = (
                "Subsidiaria extranjera nunca target — cap bajo"
                if role == "STRATEGIC"
                else "Excluido del scope"
                if role == "OUT_OF_SCOPE"
                else "—"
            )
            row = self._row(ws, row, role, f"{cap:.2f}", sig, styles)
        row += 1

        # Sección 6: Edad y fatiga
        row = write_section_header(ws, row, "Owner fatigue thresholds", styles)
        row = write_table_headers(ws, row, ["Threshold", "Valor (años)", "Efecto"], styles)
        at = thresholds.age_thresholds
        age_rows: list[tuple[str, int, str]] = [
            ("edad_madura_min", at.edad_madura_min, "Marca empresa como madura"),
            ("fatigue_age_medium", at.fatigue_age_medium, "+0.30 a fatigue score"),
            ("fatigue_age_high", at.fatigue_age_high, "+0.50 a fatigue score"),
        ]
        for label, age_val, sig in age_rows:
            row = self._row(ws, row, label, str(age_val), sig, styles)

    def _row(
        self,
        ws: Worksheet,
        row: int,
        label: str,
        value: str,
        notes: str,
        styles: StyleSet,
    ) -> int:
        ws.cell(row=row, column=1, value=label).style = styles.body_emphasis
        ws.cell(row=row, column=2, value=value).style = styles.body_centered
        ws.cell(row=row, column=3, value=notes).style = styles.body
        return row + 1
