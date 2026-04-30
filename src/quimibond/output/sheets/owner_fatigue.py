"""Hoja 6 — Owner Fatigue: score y razones detectadas por target."""

from __future__ import annotations

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from quimibond.config_loader import Config
from quimibond.models import PipelineData
from quimibond.output.helpers import (
    freeze_below,
    set_column_widths,
    write_table_headers,
    write_title,
)
from quimibond.output.styles import StyleSet, score_heatmap_fill


class OwnerFatigueSheet:
    name = "6. Owner Fatigue"

    def build(
        self,
        wb: Workbook,
        ws: Worksheet,
        data: PipelineData,
        config: Config,
        styles: StyleSet,
    ) -> None:
        set_column_widths(ws, [11, 36, 16, 11, 14, 10, 60])

        row = write_title(
            ws, 1,
            "Owner Fatigue",
            "Score 0–1: probabilidad de que el dueño esté listo para vender. "
            "Señales detectadas desde EMIS — refinar con conversaciones cualitativas.",
            styles,
        )

        # Solo targets accionables ordenados por fatigue desc
        targets = [
            c for c in data.companies
            if c.pe_role in ("PLATFORM_CANDIDATE", "PRIMARY_BOLT_ON", "TUCK_IN")
        ]
        targets.sort(key=lambda c: (-c.fatigue.score, c.emis_id))

        headers = ["EMIS ID", "Empresa", "PE Role", "Edad", "Familia", "Fatigue", "Señales"]
        row = write_table_headers(ws, row, headers, styles)
        freeze_below(ws, row - 1, col=2)

        for c in targets:
            col = 1
            ws.cell(row=row, column=col, value=c.emis_id).style = styles.body_centered
            col += 1
            ws.cell(row=row, column=col, value=c.company_name).style = styles.body
            col += 1
            ws.cell(row=row, column=col, value=c.pe_role).style = styles.body_centered
            col += 1
            ws.cell(
                row=row, column=col,
                value=c.age_years if c.age_years is not None else "—",
            ).style = styles.body_centered
            col += 1
            ws.cell(
                row=row, column=col,
                value=c.classification.detected_family or "—",
            ).style = styles.body_centered
            col += 1
            fat_cell = ws.cell(row=row, column=col, value=c.fatigue.score)
            fat_cell.style = styles.score
            fat_cell.fill = score_heatmap_fill(c.fatigue.score)
            col += 1
            ws.cell(
                row=row, column=col,
                value=" · ".join(c.fatigue.signals) or "—",
            ).style = styles.body
            ws.row_dimensions[row].height = 32
            row += 1
