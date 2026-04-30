"""Hoja 4 — Tres Palancas: detalle de cost / revenue / arbitrage por target."""

from __future__ import annotations

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from quimibond.config_loader import Config
from quimibond.models import PipelineData
from quimibond.output.helpers import (
    freeze_below,
    set_column_widths,
    write_section_header,
    write_table_headers,
    write_title,
)
from quimibond.output.styles import StyleSet, score_heatmap_fill

# Roles operacionales (excluye STRATEGIC y OUT_OF_SCOPE).
INCLUDED_ROLES = ("PLATFORM_CANDIDATE", "PRIMARY_BOLT_ON", "TUCK_IN")


class TresPalancasSheet:
    name = "4. Tres Palancas"

    def build(
        self,
        wb: Workbook,
        ws: Worksheet,
        data: PipelineData,
        config: Config,
        styles: StyleSet,
    ) -> None:
        set_column_widths(ws, [11, 32, 16, 8, 50, 8, 50, 8, 50, 10])

        row = write_title(
            ws, 1,
            "Tres palancas",
            "Cost · Revenue · Arbitrage. Justificación textual por target. "
            f"Pesos: cost={config.pe_playbook.scoring_weights.lever_cost} · "
            f"revenue={config.pe_playbook.scoring_weights.lever_revenue} · "
            f"arbitrage={config.pe_playbook.scoring_weights.lever_arbitrage}",
            styles,
        )

        targets = [c for c in data.companies if c.pe_role in INCLUDED_ROLES]
        targets.sort(key=lambda c: (-c.levers.combined, c.emis_id))

        row = write_section_header(
            ws, row, f"Targets accionables ({len(targets)})", styles,
        )

        headers = [
            "EMIS ID",
            "Empresa",
            "Rol",
            "Cost",
            "Cost — justificación",
            "Revenue",
            "Revenue — justificación",
            "Arbitrage",
            "Arbitrage — justificación",
            "Combined",
        ]
        row = write_table_headers(ws, row, headers, styles)
        freeze_below(ws, row - 1, col=2)

        for c in targets:
            col = 1
            ws.cell(row=row, column=col, value=c.emis_id).style = styles.body_centered
            col += 1
            ws.cell(row=row, column=col, value=c.company_name).style = styles.body
            col += 1
            ws.cell(row=row, column=col, value=c.pe_role).style = styles.body_centered
            col += 1

            cost_cell = ws.cell(row=row, column=col, value=c.levers.cost)
            cost_cell.style = styles.score
            cost_cell.fill = score_heatmap_fill(c.levers.cost)
            col += 1
            ws.cell(row=row, column=col, value=c.levers.cost_justification).style = styles.body
            col += 1

            rev_cell = ws.cell(row=row, column=col, value=c.levers.revenue)
            rev_cell.style = styles.score
            rev_cell.fill = score_heatmap_fill(c.levers.revenue)
            col += 1
            ws.cell(row=row, column=col, value=c.levers.revenue_justification).style = styles.body
            col += 1

            arb_cell = ws.cell(row=row, column=col, value=c.levers.arbitrage)
            arb_cell.style = styles.score
            arb_cell.fill = score_heatmap_fill(c.levers.arbitrage)
            col += 1
            ws.cell(row=row, column=col, value=c.levers.arbitrage_justification).style = styles.body
            col += 1

            comb_cell = ws.cell(row=row, column=col, value=c.levers.combined)
            comb_cell.style = styles.score
            comb_cell.fill = score_heatmap_fill(c.levers.combined)

            ws.row_dimensions[row].height = 50
            row += 1
