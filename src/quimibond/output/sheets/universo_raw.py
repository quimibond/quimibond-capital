"""Hoja 2 — Universo Raw: todas las empresas con todos los campos enriquecidos."""

from __future__ import annotations

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from quimibond.config_loader import Config
from quimibond.models import Company, PipelineData
from quimibond.output.helpers import (
    freeze_below,
    set_column_widths,
    write_table_headers,
    write_title,
)
from quimibond.output.styles import StyleSet, role_fill

HEADERS = [
    "EMIS ID",
    "Empresa",
    "RFC",
    "Estado",
    "Municipio",
    "Subsector",
    "Priority",
    "Procesos QB",
    "Cliente B2B",
    "Capital",
    "Familia",
    "Revenue (USD MM)",
    "EBITDA est. (USD MM)",
    "Empleados",
    "Edad (años)",
    "Productividad USD/emp",
    "PE Role",
    "Lever combined",
    "Fatigue",
]
COL_WIDTHS = [11, 36, 14, 18, 18, 22, 10, 24, 14, 16, 14, 14, 14, 11, 11, 16, 22, 11, 10]


class UniversoRawSheet:
    name = "2. Universo Raw"

    def build(
        self,
        wb: Workbook,
        ws: Worksheet,
        data: PipelineData,
        config: Config,
        styles: StyleSet,
    ) -> None:
        set_column_widths(ws, COL_WIDTHS)

        row = write_title(
            ws, 1,
            "Universo enriquecido",
            f"{data.n_companies} empresas con clasificación + scoring · "
            "ordenadas por EMIS ID",
            styles,
        )

        row = write_table_headers(ws, row, HEADERS, styles)
        freeze_below(ws, row - 1, col=2)

        for c in data.companies:
            self._write_row(ws, row, c, styles)
            row += 1

    def _write_row(
        self, ws: Worksheet, row: int, c: Company, styles: StyleSet,
    ) -> None:
        clf = c.classification
        col = 1
        ws.cell(row=row, column=col, value=c.emis_id).style = styles.body_centered
        col += 1
        ws.cell(row=row, column=col, value=c.company_name).style = styles.body
        col += 1
        ws.cell(row=row, column=col, value=c.rfc or "").style = styles.body_centered
        col += 1
        ws.cell(row=row, column=col, value=c.state or "").style = styles.body
        col += 1
        ws.cell(row=row, column=col, value=c.municipality or "").style = styles.body
        col += 1
        ws.cell(row=row, column=col, value=clf.subsector).style = styles.body
        col += 1
        ws.cell(row=row, column=col, value=clf.subsector_priority).style = styles.body_centered
        col += 1
        ws.cell(
            row=row, column=col,
            value=", ".join(clf.quimibond_processes) or "—",
        ).style = styles.body
        col += 1
        ws.cell(row=row, column=col, value=clf.main_client_type).style = styles.body_centered
        col += 1
        ws.cell(row=row, column=col, value=clf.capital_origin).style = styles.body_centered
        col += 1
        ws.cell(row=row, column=col, value=clf.detected_family or "").style = styles.body_centered
        col += 1

        # Revenue (USD MM, mostrado en millones literal — formato custom multiplica por 1e-6
        # del USD; aquí guardamos USD * 1e6 para que el formato '$#,##0.0,,"M"' funcione).
        if c.revenue_usd_mm is not None:
            ws.cell(row=row, column=col, value=c.revenue_usd_mm * 1_000_000).style = (
                styles.currency_usd_mm
            )
        else:
            ws.cell(row=row, column=col, value="—").style = styles.body_centered
        col += 1

        if c.ebitda_estimate_usd_mm is not None:
            ws.cell(row=row, column=col, value=c.ebitda_estimate_usd_mm * 1_000_000).style = (
                styles.currency_usd_mm
            )
        else:
            ws.cell(row=row, column=col, value="—").style = styles.body_centered
        col += 1

        if c.employees is not None:
            ws.cell(row=row, column=col, value=c.employees).style = styles.integer
        else:
            ws.cell(row=row, column=col, value="—").style = styles.body_centered
        col += 1

        if c.age_years is not None:
            ws.cell(row=row, column=col, value=c.age_years).style = styles.integer
        else:
            ws.cell(row=row, column=col, value="—").style = styles.body_centered
        col += 1

        if c.productivity_usd_per_employee is not None:
            ws.cell(row=row, column=col, value=c.productivity_usd_per_employee).style = (
                styles.integer
            )
        else:
            ws.cell(row=row, column=col, value="—").style = styles.body_centered
        col += 1

        role_cell = ws.cell(row=row, column=col, value=c.pe_role)
        role_cell.style = styles.body_centered
        role_cell.fill = role_fill(c.pe_role)
        col += 1

        ws.cell(row=row, column=col, value=c.levers.combined).style = styles.score
        col += 1
        ws.cell(row=row, column=col, value=c.fatigue.score).style = styles.score
