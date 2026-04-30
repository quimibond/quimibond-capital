"""Hoja 7 — Saturation Check: análisis de saturación por subsegmento."""

from __future__ import annotations

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from quimibond.config_loader import Config
from quimibond.models import PipelineData, SaturationVerdictType
from quimibond.output.helpers import (
    set_column_widths,
    write_table_headers,
    write_title,
)
from quimibond.output.styles import (
    COLOR_AMBER,
    COLOR_GREEN,
    COLOR_GREY_LIGHT,
    COLOR_RED,
    StyleSet,
)
from openpyxl.styles import PatternFill


VERDICT_COLORS: dict[SaturationVerdictType, str] = {
    "Atractivo": COLOR_GREEN,
    "Mixto": COLOR_AMBER,
    "Saturado": COLOR_RED,
    "Insuficiente": COLOR_GREY_LIGHT,
}


class SaturationCheckSheet:
    name = "7. Saturation Check"

    def build(
        self,
        wb: Workbook,
        ws: Worksheet,
        data: PipelineData,
        config: Config,
        styles: StyleSet,
    ) -> None:
        set_column_widths(ws, [30, 12, 14, 12, 14, 50])

        row = write_title(
            ws, 1,
            "Saturación por subsegmento",
            "Consolidados = subsidiarias extranjeras + públicas + revenue ≥ "
            f"${config.thresholds.revenue_brackets_usd_mm.platform_min:.0f}M.  "
            "Veredicto según ratio consolidados/total.",
            styles,
        )

        headers = ["Subsegmento", "Total", "Consolidadas", "Accesibles",
                   "Veredicto", "Notas"]
        row = write_table_headers(ws, row, headers, styles)

        # Orden: primero los Atractivos, luego Mixto, Saturado, Insuficiente
        order = ("Atractivo", "Mixto", "Saturado", "Insuficiente")
        verdict_order = {v: i for i, v in enumerate(order)}
        sorted_sat = sorted(
            data.saturation,
            key=lambda s: (verdict_order.get(s.verdict, 99), -s.total_companies, s.subsegment),
        )

        for s in sorted_sat:
            col = 1
            ws.cell(row=row, column=col, value=s.subsegment).style = styles.body_emphasis
            col += 1
            ws.cell(row=row, column=col, value=s.total_companies).style = styles.integer
            col += 1
            ws.cell(row=row, column=col, value=s.consolidated).style = styles.integer
            col += 1
            ws.cell(row=row, column=col, value=s.accessible).style = styles.integer
            col += 1
            verdict_cell = ws.cell(row=row, column=col, value=s.verdict)
            verdict_cell.style = styles.body_centered
            verdict_cell.fill = PatternFill(
                "solid", start_color=VERDICT_COLORS[s.verdict],
            )
            col += 1
            ws.cell(row=row, column=col, value=s.notes).style = styles.body
            ws.row_dimensions[row].height = 30
            row += 1
