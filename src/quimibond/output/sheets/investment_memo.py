"""Hoja 9 — Investment Memo: estructura para presentación a LPs externos."""

from __future__ import annotations

from collections import Counter

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from quimibond.config_loader import Config
from quimibond.models import PipelineData
from quimibond.output.helpers import (
    set_column_widths,
    write_section_header,
    write_table_headers,
    write_title,
)
from quimibond.output.styles import StyleSet


class InvestmentMemoSheet:
    name = "9. Investment Memo"

    def build(
        self,
        wb: Workbook,
        ws: Worksheet,
        data: PipelineData,
        config: Config,
        styles: StyleSet,
    ) -> None:
        set_column_widths(ws, [32, 70])

        row = write_title(
            ws, 1,
            "Investment Memo — Quimibond Capital",
            "Resumen ejecutivo del fondo y la tesis para presentación a LPs.",
            styles,
        )

        # Tesis
        row = write_section_header(ws, row, "Tesis de inversión", styles)
        thesis = (
            "Roll-up clásico: adquirir empresas familiares textiles mexicanas a múltiplos "
            "bajos (4–6× EBITDA) e integrarlas alrededor de Quimibond (operadora ancla en "
            "Toluca: tejido circular + tintorería + acabado para clientes Tier-1 automotrices "
            "como LEAR y SHAWMUT). Salida del consolidado a 8–9× EBITDA. Aprovecha "
            "nearshoring + T-MEC + transición EVs + fragmentación del sector mexicano."
        )
        row = self._narrative(ws, row, "", thesis, styles)
        row += 1

        # Tamaño del universo
        row = write_section_header(ws, row, "Universo objetivo", styles)
        n_total = data.n_companies
        n_critica = sum(1 for c in data.companies if c.classification.subsector_priority == "Crítica")
        n_familiar = sum(1 for c in data.companies if c.is_familiar_mx)
        n_targets = sum(
            1 for c in data.companies
            if c.pe_role in ("PLATFORM_CANDIDATE", "PRIMARY_BOLT_ON", "TUCK_IN")
        )
        row = self._narrative(
            ws, row, "Total empresas analizadas", f"{n_total}", styles,
        )
        row = self._narrative(
            ws, row, "Subsector crítico (no tejidos / automotriz / recubrimientos)", f"{n_critica}", styles,
        )
        row = self._narrative(
            ws, row, "Familiares MX detectadas", f"{n_familiar}", styles,
        )
        row = self._narrative(
            ws, row, "Targets accionables (Platform/Bolt-on/Tuck-in)", f"{n_targets}", styles,
        )
        row += 1

        # Asunciones del modelo
        row = write_section_header(ws, row, "Asunciones del modelo financiero", styles)
        playbook = config.pe_playbook
        row = self._narrative(
            ws, row, "EBITDA margin asumido (textil mid-market MX)",
            f"{playbook.ebitda_margin_default*100:.0f}%", styles,
        )
        row = self._narrative(
            ws, row, "Buy multiples por bracket (USD MM revenue)",
            self._brackets_summary(playbook), styles,
        )
        row = self._narrative(
            ws, row, "Exit multiple objetivo (consolidado)",
            f"{playbook.exit_multiple_default:.1f}×", styles,
        )
        row = self._narrative(
            ws, row, "Hold period",
            f"{playbook.hold_period_years} años", styles,
        )
        row = self._narrative(
            ws, row, "Discount rate corporativa",
            f"{playbook.discount_rate*100:.0f}%", styles,
        )
        row = self._narrative(
            ws, row, "Tipo de cambio asumido",
            f"{playbook.exchange_rate_mxn_usd:.0f} MXN/USD", styles,
        )
        row += 1

        # Top 5 targets por combined score
        row = write_section_header(ws, row, "Top 5 targets por combined score", styles)
        targets = [
            c for c in data.companies
            if c.pe_role in ("PLATFORM_CANDIDATE", "PRIMARY_BOLT_ON", "TUCK_IN")
        ]
        targets.sort(key=lambda c: (-c.levers.combined, c.emis_id))
        row = write_table_headers(
            ws, row,
            ["Empresa", "Detalle"],
            styles,
        )
        for c in targets[:5]:
            rev = f"${c.revenue_usd_mm:.0f}M" if c.revenue_usd_mm else "rev=?"
            emp = f"{c.employees} emp" if c.employees else "emp=?"
            detalle = (
                f"{c.pe_role} · {c.classification.subsector} · {rev} · {emp} · "
                f"familia={c.classification.detected_family or '—'} · "
                f"combined={c.levers.combined:.2f}"
            )
            ws.cell(row=row, column=1, value=c.company_name).style = styles.body_emphasis
            ws.cell(row=row, column=2, value=detalle).style = styles.body
            ws.row_dimensions[row].height = 30
            row += 1
        row += 1

        # Saturación
        row = write_section_header(ws, row, "Saturación de subsectores críticos", styles)
        critical_sats = [s for s in data.saturation if s.verdict in ("Atractivo", "Mixto")]
        for s in sorted(critical_sats, key=lambda x: -x.total_companies)[:6]:
            line = (
                f"{s.subsegment}: {s.total_companies} empresas, "
                f"{s.consolidated} consolidadas → {s.verdict}"
            )
            ws.cell(row=row, column=1, value=s.subsegment).style = styles.body_emphasis
            ws.cell(row=row, column=2, value=line).style = styles.body
            row += 1

    def _narrative(
        self,
        ws: Worksheet,
        row: int,
        label: str,
        text: str,
        styles: StyleSet,
    ) -> int:
        if label:
            ws.cell(row=row, column=1, value=label).style = styles.body_emphasis
        ws.cell(row=row, column=2, value=text).style = styles.body
        ws.row_dimensions[row].height = max(28, 18 * (1 + len(text) // 100))
        return row + 1

    def _brackets_summary(self, playbook) -> str:  # type: ignore[no-untyped-def]
        parts = []
        prev = 0.0
        for b in playbook.buy_multiples_by_size:
            if b.max_revenue is None:
                parts.append(f">${prev:.0f}M: {b.multiple:.1f}×")
            else:
                parts.append(f"${prev:.0f}–${b.max_revenue:.0f}M: {b.multiple:.1f}×")
                prev = b.max_revenue
        return " · ".join(parts)
