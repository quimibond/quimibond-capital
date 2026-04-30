"""
Helpers compartidos por los builders de hojas.

Mantienen la firma corta (no inyectan estilos arbitrarios). Cualquier helper
de presentación que aparezca >2 veces en sheets/ debe vivir aquí.
"""

from __future__ import annotations

from collections.abc import Iterable
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from quimibond.output.styles import StyleSet


def write_title(ws: Worksheet, row: int, title: str, subtitle: str, styles: StyleSet) -> int:
    """Escribe título y subtítulo. Devuelve la próxima fila libre."""
    ws.cell(row=row, column=1, value=title).style = styles.title
    ws.row_dimensions[row].height = 30
    ws.cell(row=row + 1, column=1, value=subtitle).style = styles.subtitle
    ws.row_dimensions[row + 1].height = 20
    return row + 3  # deja un renglón en blanco


def write_section_header(ws: Worksheet, row: int, text: str, styles: StyleSet) -> int:
    """Header de sección menor. Devuelve la próxima fila libre."""
    ws.cell(row=row, column=1, value=text).style = styles.section_header
    ws.row_dimensions[row].height = 22
    return row + 1


def write_table_headers(
    ws: Worksheet,
    row: int,
    headers: Iterable[str],
    styles: StyleSet,
    col_offset: int = 0,
) -> int:
    """Escribe headers de tabla. Devuelve next row."""
    for i, h in enumerate(headers, start=1 + col_offset):
        ws.cell(row=row, column=i, value=h).style = styles.table_header
    ws.row_dimensions[row].height = 32
    return row + 1


def set_column_widths(ws: Worksheet, widths: Iterable[int]) -> None:
    """Set widths en orden A, B, C, ..."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def freeze_below(ws: Worksheet, row: int, col: int = 1) -> None:
    """Freeze panes debajo de `row` (1-based) y a la derecha de col."""
    ws.freeze_panes = ws.cell(row=row + 1, column=col + 1).coordinate


def write_kv_pair(
    ws: Worksheet,
    row: int,
    label: str,
    value: Any,
    styles: StyleSet,
    *,
    value_style: str | None = None,
    label_col: int = 1,
    value_col: int = 2,
) -> int:
    """Escribe un par label/value. Devuelve next row."""
    ws.cell(row=row, column=label_col, value=label).style = styles.body_emphasis
    cell = ws.cell(row=row, column=value_col, value=value)
    cell.style = value_style or styles.body_centered
    return row + 1
